from playwright.sync_api import sync_playwright
import time
import os
import argparse
import re
from pathlib import Path
import sys
import openpyxl
from openpyxl.cell.rich_text import CellRichText, TextBlock
from openpyxl.cell.text import InlineFont
from openpyxl.cell.cell import MergedCell

# Configuration
ROOT_DIR = Path(__file__).resolve().parent

DEFAULT_EXCEL_CANDIDATES = [
    str(ROOT_DIR / "IT23665866_Assignment 1 - Test cases.xlsx"),
    str(ROOT_DIR / "Assignment 1 - Test cases.xlsx"),
]


DEFAULT_SHEET_NAME = "Test cases"

DEFAULT_FRONTEND_URL = os.getenv("FRONTEND_URL", "https://www.pixelssuite.com/chat-translator")
DEFAULT_PANEL = "chat-sinhala"

DEFAULT_INPUT_COLUMN_CANDIDATES = [
    "Input","Input Column","Singlish Input","Test Input","Singlish","Source","Sentence","Text",
]

DEFAULT_EXPECTED_COLUMN_CANDIDATES = [
    "Sinhala","Expected_Output","Expected Output","Expected output","Expected","Expected Sinhala",
]

DEFAULT_ACTUAL_COLUMN_CANDIDATES = [
    "Actual_Output","Actual Output","Actual output","Actual",
]

DEFAULT_STATUS_COLUMN_CANDIDATES = [
    "Status","Result","Pass/Fail","Pass Fail",
]

EVIDENCE_RATIONALE_COLUMN_CANDIDATES = [
    "Evidence or Rationale for the Input Type Covered",
    "Evidence/Rationale",
    "Evidence or Rationale",
]

DEFAULT_WAIT_MS = 8000
DEFAULT_RETRIES = 8
DEFAULT_RETRY_WAIT_MS = 1000
DEFAULT_TYPE_DELAY_MS = 50
DEFAULT_TIMEOUT_MS = 60000
DEFAULT_SLOW_MO_MS = 300
DEFAULT_SAVE_EVERY = 1



def _configure_stdout():
    try:
        sys.stdout.reconfigure(encoding="utf-8", errors="backslashreplace")
    except:
        pass


def _pick_existing_path(candidates):
    for p in candidates:
        if p and os.path.exists(p):
            return p
    return candidates[0] if candidates else None


def _resolve_path(p):
    if not p:
        return None
    path = Path(p)
    return str(path if path.is_absolute() else (ROOT_DIR / path).resolve())


def _parse_args():
    parser = argparse.ArgumentParser(
        description="Run Chat Sinhala transliteration tests from an Excel sheet."
    )
    parser.add_argument("--excel", default=None, help="Path to the Excel test case file.")
    parser.add_argument("--url", default=DEFAULT_FRONTEND_URL, help="Frontend URL to test.")
    parser.add_argument(
        "--panel",
        choices=[DEFAULT_PANEL],
        default=DEFAULT_PANEL,
        help="PixelsSuite panel to test. This assignment is scoped to Chat Sinhala only.",
    )
    parser.add_argument("--wait-ms", type=int, default=DEFAULT_WAIT_MS, help="Wait time after each test action.")
    parser.add_argument("--type-delay-ms", type=int, default=DEFAULT_TYPE_DELAY_MS, help="Delay between typed characters.")
    parser.add_argument("--slow-mo-ms", type=int, default=DEFAULT_SLOW_MO_MS, help="Playwright slow motion delay.")
    parser.add_argument("--save-every", type=int, default=DEFAULT_SAVE_EVERY, help="Save workbook every N processed rows.")
    parser.add_argument("--keep-open", action="store_true", help="Keep browser open after tests finish.")
    return parser.parse_args()


def _normalize_header(value):
    if value is None:
        return ""
    return re.sub(r"[^a-z0-9]+", "", str(value).strip().lower())


def _apply_evidence_rationale_bold(ws, header_row):
    headers = _header_values(ws, header_row)
    normalized_headers = [_normalize_header(v) for v in headers]
    evidence_col = None

    for candidate in EVIDENCE_RATIONALE_COLUMN_CANDIDATES:
        normalized_candidate = _normalize_header(candidate)
        if normalized_candidate in normalized_headers:
            evidence_col = normalized_headers.index(normalized_candidate) + 1
            break

    if not evidence_col:
        return

    bold_font = InlineFont(b=True)
    prefixes = ("Evidence:", "Rationale:")

    for row in range(header_row + 1, ws.max_row + 1):
        cell = ws.cell(row=row, column=evidence_col)
        value = cell.value
        if value is None or isinstance(value, CellRichText):
            continue

        text = str(value)
        for prefix in prefixes:
            if text.startswith(prefix):
                cell.value = CellRichText([
                    TextBlock(bold_font, prefix),
                    text[len(prefix):],
                ])
                break


def _header_values(ws, row):
    return [ws.cell(row=row, column=c).value for c in range(1, ws.max_column + 1)]


def _get_worksheet(wb, sheet_name):
    normalized_sheet_name = _normalize_header(sheet_name)
    for name in wb.sheetnames:
        if _normalize_header(name) == normalized_sheet_name:
            return wb[name]
    return wb.active



def _wait_for_output(output_locator, previous_output=""):
    deadline = time.time() + 30
    while time.time() < deadline:
        current = _read_textarea_value(output_locator)
        if current and current != previous_output:
            return True
        time.sleep(0.5)
    return False


def _first_visible(locators, timeout=1000):
    for locator in locators:
        try:
            if locator.count() > 0 and locator.first.is_visible(timeout=timeout):
                return locator.first
        except:
            continue
    return None


def _find_transliteration_controls(page):
    page.wait_for_selector("textarea", timeout=DEFAULT_TIMEOUT_MS)
    textarea_count = page.locator("textarea").count()

    input_box = None
    try:
        singlish_box = page.locator('textarea[placeholder*="Singlish" i]').first
        if singlish_box.count() > 0:
            input_box = singlish_box
    except:
        pass
    if not input_box and textarea_count >= 1:
        input_box = page.locator("textarea").nth(0)

    output_box = None
    if textarea_count >= 2:
        output_box = page.locator("textarea").nth(1)
    if not output_box:
        output_box = _first_visible([
            page.locator("div.whitespace-pre-wrap.overflow-y-auto.bg-slate-50"),
            page.locator("div.whitespace-pre-wrap").filter(has_not_text=re.compile(r"characters$", re.I)),
            page.locator('[contenteditable="true"]').nth(1),
        ])

    button = _first_visible([
        page.get_by_role("button", name=re.compile(r"^Translate$", re.I)),
        page.get_by_role("button", name=re.compile(r"Transliterate", re.I)),
        page.locator("button").filter(has_text=re.compile(r"Translate|Transliterate", re.I)),
    ])

    if not input_box:
        raise RuntimeError("Could not find the Singlish input box.")
    if not output_box:
        raise RuntimeError(f"Could not find the Sinhala output box. Found {textarea_count} textarea(s).")
    if not button:
        raise RuntimeError("Could not find the Translate button.")

    return input_box, output_box, button


def _read_textarea_value(locator):
    try:
        return (locator.input_value() or "").strip()
    except:
        try:
            return (locator.text_content() or "").strip()
        except:
            return ""


def _type_text(locator, text, delay_ms):
    locator.scroll_into_view_if_needed(timeout=3000)
    locator.click(timeout=3000)
    locator.fill("")
    locator.type(str(text), delay=delay_ms)


def _select_sinhala_panel(page, panel):
    target_text = "Chat Sinhala"

    try:
        if page.get_by_text(target_text, exact=True).first.is_visible(timeout=1000):
            page.get_by_text(target_text, exact=True).first.click(timeout=3000)
            page.wait_for_timeout(1000)
            return
    except:
        pass

    menu_triggers = [
        page.get_by_role("button", name=re.compile("Transliteration|Standard Sinhala|Chat Sinhala", re.I)).first,
        page.get_by_text(re.compile("Transliteration", re.I)).first,
        page.locator("button, [role=button], a").filter(
            has_text=re.compile("Transliteration|Standard Sinhala|Chat Sinhala", re.I)
        ).first,
    ]

    for trigger in menu_triggers:
        try:
            if trigger.is_visible(timeout=1000):
                trigger.click(timeout=3000)
                page.get_by_text(target_text, exact=True).first.click(timeout=5000)
                page.wait_for_timeout(1000)
                print(f"Selected panel: {target_text}")
                return
        except:
            continue

    print(f"Warning: Could not select '{target_text}'. Continuing with the currently open panel.")



def run_test():
    _configure_stdout()
    args = _parse_args()

    excel_path = _resolve_path(args.excel) if args.excel else _resolve_path(_pick_existing_path(DEFAULT_EXCEL_CANDIDATES))

    if not os.path.exists(excel_path):
        print(f"Excel file not found: {excel_path}")
        return

    wb = openpyxl.load_workbook(excel_path, rich_text=True)
    ws = _get_worksheet(wb, DEFAULT_SHEET_NAME)

    header_row = 1
    headers = _header_values(ws, header_row)

    def find_col(names):
        norm = [_normalize_header(v) for v in headers]
        for n in names:
            if _normalize_header(n) in norm:
                return norm.index(_normalize_header(n)) + 1
        return None

    input_col = find_col(DEFAULT_INPUT_COLUMN_CANDIDATES)
    expected_col = find_col(DEFAULT_EXPECTED_COLUMN_CANDIDATES)
    actual_col = find_col(DEFAULT_ACTUAL_COLUMN_CANDIDATES) or ws.max_column + 1
    status_col = find_col(DEFAULT_STATUS_COLUMN_CANDIDATES) or ws.max_column + 2

    ws.cell(row=header_row, column=actual_col).value = "Actual Output"
    ws.cell(row=header_row, column=status_col).value = "Status"
    _apply_evidence_rationale_bold(ws, header_row)

    print("Starting Chat Sinhala transliteration tests...")

    passed = 0
    failed = 0
    processed = 0

    with sync_playwright() as p:
        browser = p.chromium.launch(headless=False, slow_mo=args.slow_mo_ms)
        page = browser.new_page()

        page.goto(args.url, timeout=DEFAULT_TIMEOUT_MS)
        _select_sinhala_panel(page, args.panel)
        page.wait_for_selector("textarea")
        input_box, output_box, button = _find_transliteration_controls(page)

        for row in range(header_row + 1, ws.max_row + 1):
            text = ws.cell(row=row, column=input_col).value
            if not text:
                continue

            expected = ws.cell(row=row, column=expected_col).value
            expected = str(expected).strip() if expected else ""

            print(f"Row {row}: {text}")

            try:
                prev = _read_textarea_value(output_box)

                _type_text(input_box, text, args.type_delay_ms)
                button.click()

                _wait_for_output(output_box, prev)

                time.sleep(args.wait_ms / 1000)

                actual = _read_textarea_value(output_box)

                ws.cell(row=row, column=actual_col).value = actual

               
                if expected:
                    status = "PASS" if actual.strip() == expected.strip() else "FAIL"
                else:
                    status = "COLLECTED"

                ws.cell(row=row, column=status_col).value = status

                if status == "PASS":
                    passed += 1
                elif status == "FAIL":
                    failed += 1

                processed += 1
                print(f" -> {status}")

                if args.save_every > 0 and processed % args.save_every == 0:
                    wb.save(excel_path)

            except Exception as e:
                print(f"Error: {e}")
                ws.cell(row=row, column=status_col).value = "ERROR"
                wb.save(excel_path)

        if args.keep_open:
            print("Browser kept open. Press Ctrl+C to close this script when done.")
            try:
                while True:
                    time.sleep(1)
            except KeyboardInterrupt:
                pass

        browser.close()

   
    print("\n=== TEST SUMMARY ===")
    print(f"Total: {processed}")
    print(f"Passed: {passed}")
    print(f"Failed: {failed}")

    wb.save(excel_path)
    print("Done.")


if __name__ == "__main__":
    run_test()
